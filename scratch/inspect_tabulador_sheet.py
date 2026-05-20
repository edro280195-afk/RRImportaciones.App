import openpyxl

wb = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=False)
sheet = wb["TABULADOR"]

print(f"Sheet TABULADOR size: {sheet.max_row} rows x {sheet.max_column} columns")

# Print first 40 rows and first 15 columns
for r in range(1, min(100, sheet.max_row + 1)):
    row_vals = []
    for c in range(1, min(20, sheet.max_column + 1)):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        # Format formula or value nicely
        if val is not None:
            row_vals.append(f"C{c}:{val}")
    if row_vals:
        print(f"Row {r:02d}: " + " | ".join(row_vals))
