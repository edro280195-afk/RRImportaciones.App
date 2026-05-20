import openpyxl

wb_f = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=False)
wb_v = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=True)

sheet_f = wb_f["AMPARO"]
sheet_v = wb_v["AMPARO"]

print(f"AMPARO sheet rows:")
for r in range(1, sheet_f.max_row + 1):
    row_str = f"Row {r:02d}: "
    for c in range(1, sheet_f.max_column + 1):
        cell_f = sheet_f.cell(row=r, column=c)
        cell_v = sheet_v.cell(row=r, column=c)
        val_f = cell_f.value
        val_v = cell_v.value
        if val_f is not None or val_v is not None:
            col_letter = openpyxl.utils.get_column_letter(c)
            row_str += f"{col_letter}: [{val_f} -> {val_v}] | "
    print(row_str)
