import openpyxl

wb = openpyxl.load_workbook("TABULADOR 2026.xlsx", read_only=True)
sheet_name = "TRACTORES DE CARRETERA-IGNORAR"
ws = wb[sheet_name]
print(f"--- Structure of {sheet_name} ---")
for r in range(1, 15):
    row_str = f"Row {r:02d}: "
    for c in range(1, min(15, ws.max_column + 1)):
        val = ws.cell(row=r, column=c).value
        if val is not None:
            row_str += f"C{c}:{val} | "
    if any(ws.cell(row=r, column=col).value is not None for col in range(1, ws.max_column + 1)):
        print(row_str)
