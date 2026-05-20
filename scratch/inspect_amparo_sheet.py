import openpyxl

wb_f = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=False)
wb_v = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=True)

sheet_f = wb_f["AMPARO"]
sheet_v = wb_v["AMPARO"]

print(f"Sheet AMPARO size: {sheet_f.max_row} rows x {sheet_f.max_column} columns")

for r in range(1, min(100, sheet_f.max_row + 1)):
    row_str = f"Row {r:02d}: "
    for col_idx in range(1, min(10, sheet_f.max_column + 1)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell_f = sheet_f.cell(row=r, column=col_idx)
        cell_v = sheet_v.cell(row=r, column=col_idx)
        val_f = cell_f.value
        val_v = cell_v.value
        if val_f is not None or val_v is not None:
            row_str += f"{col_letter}: [{val_f} -> {val_v}] | "
    if any(sheet_f.cell(row=r, column=col_idx).value is not None for col_idx in range(1, sheet_f.max_column + 1)):
        print(row_str)
