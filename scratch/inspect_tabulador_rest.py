import openpyxl

wb = openpyxl.load_workbook("TABULADOR 2026.xlsx", data_only=False)
sheet = wb["TABULADOR"]

print("Inspecting TABULADOR rows 30 to 97:")
for r in range(30, sheet.max_row + 1):
    row_str = f"Row {r:02d}: "
    for col_idx in range(1, sheet.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell = sheet.cell(row=r, column=col_idx)
        val = cell.value
        if val is not None:
            row_str += f"{col_letter}: {val} | "
    if any(sheet.cell(row=r, column=c).value is not None for c in range(1, sheet.max_column + 1)):
        print(row_str)
