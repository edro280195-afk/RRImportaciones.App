import openpyxl

wb = openpyxl.load_workbook("TABULADOR 2026.xlsx", read_only=True)

sheets = ["AUT 1.0 A 1.5", "PICK UP\u00b4S", "AMPARO"]
for sheet_name in wb.sheetnames:
    if sheet_name in sheets or sheet_name == "PICK UP\u00b4S" or "PICK" in sheet_name:
        print(f"\n--- Structure of {sheet_name} ---")
        ws = wb[sheet_name]
        # print first 6 rows and first 14 columns
        for r in range(1, min(10, ws.max_row + 1)):
            row_str = f"Row {r:02d}: "
            for c in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_str += f"C{c}:{val} | "
            if any(ws.cell(row=r, column=col).value is not None for col in range(1, ws.max_column + 1)):
                print(row_str)
